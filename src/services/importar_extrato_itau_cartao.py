from __future__ import annotations

import datetime as _datetime_mod
import os
import re

import openpyxl
import xlrd

# Linhas de cabecalho/resumo da fatura que nunca sao lancamento de compra
# (mesmo filtro do script legado importar_extrato.py) -- comparadas em
# minusculas.
_DESCRICOES_IGNORADAS = {
    "lançamento",
    "data",
    "logotipo itaú",
    "",
    "lançamentos nacionais",
    "lançamentos internacionais",
    "encargos e serviços",
    "total",
    "subtotal",
    "emitido",
    "melhor",
    "compras feitas",
    "dólar de conversão",
    "valor em dólar",
}


def _conta_pelo_nome_arquivo(caminho_arquivo: str) -> str:
    """Identifica o cartao pelos 4 digitos finais que o Itau usa no nome do
    arquivo exportado (ex.: 9073, 2486, 1035) -- generico o bastante para
    cartoes novos, sem precisar hardcodar cada numero (validado com dado
    real: assets/finalcial/Financeiro/extrato/1035-.../*.xls)."""
    nome = os.path.basename(caminho_arquivo)
    match = re.search(r"(\d{4})", nome)
    if match:
        return f"Itaú_{match.group(1)}"
    return "Itaú_CC_Cartão"


def _data_iso(valor_bruto) -> str | None:
    """Converte a data da celula do XLS (numero serial do Excel ou texto
    dd/mm/aaaa) para ISO -- mesma logica de para_iso() do script legado."""
    if isinstance(valor_bruto, float) and valor_bruto > 40000:
        try:
            data = xlrd.xldate_as_datetime(valor_bruto, 0)
            return data.strftime("%Y-%m-%d")
        except Exception:
            return None
    texto = str(valor_bruto).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            from datetime import datetime

            return datetime.strptime(texto, formato).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parsear(caminho_arquivo: str) -> list[dict]:
    """Le uma fatura de cartao Itau (.xls legado ou .xlsx 'Fatura Paga',
    feature 011 US3 -- o Itau passou a exportar tambem nesse segundo
    formato, achado processando um extrato novo real) e retorna uma lista
    de registros no mesmo formato aceito por
    importar_historico_extrato.processar_transacoes. Despacha pelo formato
    do arquivo; a logica de cada formato fica em funcao propria."""
    if caminho_arquivo.lower().endswith(".xlsx"):
        return _parsear_fatura_paga_xlsx(caminho_arquivo)
    return _parsear_fatura_xls(caminho_arquivo)


def _parsear_fatura_xls(caminho_arquivo: str) -> list[dict]:
    """Le uma fatura de cartao Itau em .xls e retorna uma lista de
    registros no mesmo formato aceito por
    importar_historico_extrato.processar_transacoes (data, descricao,
    valor_raw, conta, fonte) -- reaproveita a mesma persistencia/
    classificacao/reconciliacao da migracao historica (research.md #10/#11,
    T048), sem duplicar essa logica aqui. Porta os mesmos filtros de
    cabecalho/total/"pagamento efetuado" do script legado."""
    workbook = xlrd.open_workbook(caminho_arquivo)
    planilha = workbook.sheet_by_index(0)
    conta = _conta_pelo_nome_arquivo(caminho_arquivo)
    fonte = os.path.basename(caminho_arquivo)

    registros: list[dict] = []
    for linha in range(planilha.nrows):
        celulas = [planilha.cell_value(linha, coluna) for coluna in range(planilha.ncols)]
        if len(celulas) < 3:
            continue

        data_bruta = celulas[0]
        descricao = str(celulas[1]).strip() if celulas[1] else ""
        valor_bruto = celulas[3] if len(celulas) > 3 else None

        if not descricao or descricao.lower() in _DESCRICOES_IGNORADAS:
            continue
        if str(data_bruta).strip() in ("", "data", "lançamento"):
            continue
        if not isinstance(valor_bruto, (int, float)):
            continue
        if "total" in descricao.lower() or "total" in str(data_bruta).lower():
            continue
        if "pagamento efetuado" in descricao.lower():
            continue

        data_iso = _data_iso(data_bruta)
        if not data_iso:
            continue

        registros.append(
            {
                "data": data_iso,
                "descricao": descricao,
                "valor_raw": float(valor_bruto),
                "conta": conta,
                "fonte": fonte,
                # Todo cartao Itau importado por este parser e do Marcelo --
                # titular fixo, mesmo espirito do parser do extrato BB
                # (feature 011, research.md #7).
                "titular": "marcelo",
            }
        )

    return registros


def _parsear_fatura_paga_xlsx(caminho_arquivo: str) -> list[dict]:
    """Le uma fatura de cartao Itau no formato novo 'Fatura Paga' (.xlsx):
    metadados no topo (Nome/Agencia/Conta/resumo do cartao), depois uma
    tabela com colunas Data/Lancamento/Parcelamento/Valor/.../Titularidade.
    Diferente do .xls legado, a data ja vem como datetime real e o valor
    ja vem como float (nao precisa de conversao de serial do Excel nem de
    formato BR de decimal). So a linha com data valida e valor numerico e
    uma transacao de verdade -- todo o resto (metadados, cabecalho,
    'Subtotal', texto de rodape) fica sem data valida e cai fora sozinho,
    sem precisar de lista de exclusao por texto (research.md da feature
    011, achado processando um extrato real do usuario)."""
    workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    planilha = workbook[workbook.sheetnames[0]]
    conta = _conta_pelo_nome_arquivo(caminho_arquivo)
    fonte = os.path.basename(caminho_arquivo)

    registros: list[dict] = []
    for linha in planilha.iter_rows(min_row=1, values_only=True):
        if len(linha) < 5:
            continue
        data_bruta, descricao_bruta, _parcelamento, valor_bruto = linha[1], linha[2], linha[3], linha[4]

        if not isinstance(data_bruta, _datetime_mod.datetime):
            continue
        if not isinstance(valor_bruto, (int, float)):
            continue
        descricao = str(descricao_bruta).strip() if descricao_bruta else ""
        if not descricao:
            continue
        # Mesmo filtro do .xls legado: o pagamento da fatura ja e contado
        # do lado da conta corrente (INT ITAU MULT/CLICK) -- incluir aqui
        # de novo duplicaria o evento (conta diferente, nao deduplica por
        # fingerprint).
        if "pagamento efetuado" in descricao.lower():
            continue

        registros.append(
            {
                "data": data_bruta.strftime("%Y-%m-%d"),
                "descricao": descricao,
                "valor_raw": float(valor_bruto),
                "conta": conta,
                "fonte": fonte,
                "titular": "marcelo",
            }
        )

    return registros
