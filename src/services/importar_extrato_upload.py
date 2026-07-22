from __future__ import annotations

import openpyxl
import xlrd

from src.services import (
    importar_extrato_bb,
    importar_extrato_itau_cartao,
    importar_extrato_itau_cc,
    importar_fatura_mercado_pago,
)

_MAX_LINHAS_INSPECIONADAS = 20


class FormatoNaoReconhecidoError(Exception):
    """Extensao suportada mas nenhuma assinatura de conteudo conhecida
    bateu, ou extensao fora das suportadas -- aborta sem adivinhar
    (Principio VII: degradar significa recusar com seguranca, nao
    interpretar com o parser errado -- research.md #3)."""


def _linhas_xls(caminho: str) -> list[list]:
    workbook = xlrd.open_workbook(caminho)
    planilha = workbook.sheet_by_index(0)
    limite = min(_MAX_LINHAS_INSPECIONADAS, planilha.nrows)
    return [[planilha.cell_value(linha, coluna) for coluna in range(planilha.ncols)] for linha in range(limite)]


def _linhas_xlsx(caminho: str) -> list[tuple]:
    workbook = openpyxl.load_workbook(caminho, data_only=True)
    planilha = workbook[workbook.sheetnames[0]]
    return list(planilha.iter_rows(max_row=_MAX_LINHAS_INSPECIONADAS, values_only=True))


def _contem_texto(linhas, *alvos: str) -> bool:
    """True se QUALQUER uma das celulas nas linhas inspecionadas contiver
    (substring, case-insensitive) qualquer um dos textos-alvo."""
    alvos_normalizados = [alvo.strip().lower() for alvo in alvos]
    for linha in linhas:
        for celula in linha:
            if not celula:
                continue
            texto_celula = str(celula).strip().lower()
            if any(alvo in texto_celula for alvo in alvos_normalizados):
                return True
    return False


def _detectar_xls(caminho: str) -> str:
    """Distingue os 2 formatos que usam .xls (research.md #1): extrato de
    conta corrente Itau tem coluna "saldos" (unica entre os dois); fatura
    de cartao Itau legada nao tem essa coluna, so "lancamento" + "valor"."""
    linhas = _linhas_xls(caminho)
    if _contem_texto(linhas, "saldos"):
        return "itau_extrato_cc"
    if _contem_texto(linhas, "lançamento") and _contem_texto(linhas, "valor"):
        return "itau_fatura_cartao"
    raise FormatoNaoReconhecidoError(
        "Não foi possível reconhecer o formato deste arquivo .xls. "
        "Formatos aceitos: fatura de cartão Itaú, extrato de conta corrente Itaú."
    )


def _detectar_xlsx(caminho: str) -> str:
    """Distingue os 2 formatos que usam .xlsx (research.md #1): fatura
    Itau nova ("Fatura Paga") tem esse texto ou a coluna "Titularidade";
    extrato BB tem as colunas "Detalhes"/"N° documento"/"Tipo Lançamento"."""
    linhas = _linhas_xlsx(caminho)
    if _contem_texto(linhas, "fatura paga", "titularidade"):
        return "itau_fatura_cartao"
    if _contem_texto(linhas, "tipo lançamento", "n° documento"):
        return "bb_extrato_cc"
    raise FormatoNaoReconhecidoError(
        "Não foi possível reconhecer o formato deste arquivo .xlsx. "
        "Formatos aceitos: fatura de cartão Itaú (novo formato), extrato BB."
    )


def detectar_e_parsear(caminho_arquivo: str, nome_original: str) -> tuple[str, list[dict]]:
    """Detecta qual dos 4 formatos suportados o arquivo enviado e, sem
    duplicar nenhuma logica de parsing, despacha para o `parsear()` ja
    existente daquele formato (research.md #2). Retorna (formato
    detectado, registros) -- o chamador (rota HTTP) usa o formato so para
    informar o usuario, os registros vao direto pra
    `processar_transacoes`."""
    extensao = nome_original.rsplit(".", 1)[-1].lower() if "." in nome_original else ""

    if extensao == "pdf":
        # Unico formato suportado nessa extensao -- sem ambiguidade a
        # resolver aqui; o proprio parser valida o conteudo ("Emitida
        # em") e levanta FaturaInvalidaError se nao for uma fatura
        # Mercado Pago de verdade.
        return "mercado_pago_fatura", importar_fatura_mercado_pago.parsear(caminho_arquivo)

    if extensao == "xls":
        formato = _detectar_xls(caminho_arquivo)
    elif extensao == "xlsx":
        formato = _detectar_xlsx(caminho_arquivo)
    else:
        raise FormatoNaoReconhecidoError(
            f"Extensão '.{extensao}' não suportada. Formatos aceitos: "
            "fatura Itaú (.xls/.xlsx), extrato de conta corrente Itaú (.xls), "
            "extrato BB (.xlsx), fatura Mercado Pago (.pdf)."
        )

    if formato == "itau_extrato_cc":
        return formato, importar_extrato_itau_cc.parsear(caminho_arquivo)
    if formato == "itau_fatura_cartao":
        return formato, importar_extrato_itau_cartao.parsear(caminho_arquivo)
    return formato, importar_extrato_bb.parsear(caminho_arquivo)
