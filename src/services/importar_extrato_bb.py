from __future__ import annotations

import os
import re
from datetime import datetime

import openpyxl

# "Saldo Anterior"/"Saldo do dia"/"S A L D O" (ultima linha do arquivo, com
# letras espacadas) sao linhas de conferencia de saldo do extrato do BB,
# nunca movimentacao real (research.md #6) -- comparado sem espaco e em
# minusculas pra cobrir a variante com letras espacadas.
_LANCAMENTOS_IGNORADOS = {"saldoanterior", "saldododia", "saldo"}


def _eh_linha_de_saldo(lancamento_texto: str) -> bool:
    normalizado = re.sub(r"\s+", "", lancamento_texto).lower()
    return normalizado in _LANCAMENTOS_IGNORADOS

# "Detalhes" do extrato do BB costuma vir prefixado com "dd/mm hh:mm " antes
# do nome do estabelecimento/contraparte (ex.: "02/01 13:50 INSTITUTO CG
# CLIN ODONTOL") -- ruido pra classificacao/fingerprint (research.md #5).
_RE_PREFIXO_TIMESTAMP = re.compile(r"^\d{2}/\d{2} \d{2}:\d{2}\s+")

# Todo arquivo desta pasta pertence a mesma e unica conta corrente da
# Cristine -- titular fixo, sem heuristica (research.md #7).
TITULAR = "cristine"
CONTA = "BB_Cristine"


def _data_iso(valor_bruto: str) -> str | None:
    texto = str(valor_bruto).strip()
    if not texto or texto == "00/00/0000":
        return None
    try:
        return datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _valor_float(valor_bruto: str) -> float | None:
    """Extrato do BB traz o valor como texto em formato BR (milhar com
    ponto, decimal com virgula, ex.: '-1.234,56')."""
    texto = str(valor_bruto).strip().replace(".", "").replace(",", ".")
    if not texto:
        return None
    try:
        return float(texto)
    except ValueError:
        return None


def _descricao(lancamento: str, detalhes: str) -> str:
    """Concatena Lancamento + Detalhes -- nenhum dos dois sozinho e
    confiavel como descricao (validado com os 5 arquivos reais): Detalhes
    costuma ter o nome do estabelecimento/contraparte ('Pix - Enviado' +
    'INSTITUTO CG CLIN ODONTOL'), mas as vezes e generico e quem carrega a
    informacao util e o Lancamento ('Tarifa Pacote de Servicos' + 'Cobranca
    referente 06/01/2026')."""
    lancamento = lancamento.strip()
    detalhes = detalhes.strip()
    if detalhes:
        detalhes = _RE_PREFIXO_TIMESTAMP.sub("", detalhes).strip()
    if detalhes:
        return f"{lancamento} - {detalhes}" if lancamento else detalhes
    return lancamento


def parsear(caminho_arquivo: str) -> list[dict]:
    """Le um extrato de conta corrente do Banco do Brasil (.xlsx, colunas
    Data/Lancamento/Detalhes/N. documento/Valor/Tipo Lancamento) e retorna
    uma lista de registros no mesmo formato aceito por
    importar_historico_extrato.processar_transacoes (data, descricao,
    valor_raw, conta, fonte, titular) -- reaproveita a mesma persistencia/
    classificacao/reconciliacao da migracao historica (research.md #2), sem
    duplicar essa logica aqui."""
    workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    planilha = workbook[workbook.sheetnames[0]]
    fonte = os.path.basename(caminho_arquivo)

    registros: list[dict] = []
    for linha in planilha.iter_rows(min_row=2, values_only=True):
        if len(linha) < 5:
            continue
        data_bruta, lancamento, detalhes, _num_documento, valor_bruto = linha[:5]

        lancamento_texto = str(lancamento).strip() if lancamento else ""
        if _eh_linha_de_saldo(lancamento_texto):
            continue

        data_iso = _data_iso(data_bruta) if data_bruta else None
        if not data_iso:
            continue

        valor = _valor_float(valor_bruto) if valor_bruto else None
        if valor is None:
            continue

        descricao = _descricao(lancamento_texto, str(detalhes).strip() if detalhes else "")
        if not descricao:
            continue

        registros.append(
            {
                "data": data_iso,
                "descricao": descricao,
                "valor_raw": valor,
                "conta": CONTA,
                "fonte": fonte,
                "titular": TITULAR,
            }
        )

    return registros
